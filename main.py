import os
# import time
import re
import requests
import openpyxl
from bs4 import BeautifulSoup
# import ssl
# ssl._create_default_https_context = ssl._create_unverified_context


def get_html(post_name, tab=None, pn=None):
    """
    获取html
    :param post_name: 贴吧名
    :param tab: 标签名
    :param pn: 页码
    :return:
    """
    try:
        url = 'https://tieba.baidu.com/f'

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) \
                          Chrome/75.0.3770.100 Safari/537.36'
        }
        # 标签栏选择:
        # 核心区：corearea; 看帖：main
        data = {
            'kw': post_name,
            'tab': tab,
            'pn': pn,
        }
        response = requests.get(url, params=data, headers=headers, timeout=30)
        # 必须修改HTML页面，把HTML结束标签改到最后，否则soup解析只到原来的HTML标签就结束了，后面的code标签里的内容被丢弃
        html = response.text.replace('</body>', '')
        html = html.replace('</html>', '')
        response = html + '</body></html>'
        return response
    except RuntimeError:
        return 'ERROR'


def get_post_info(html, pn, keywords):
    """
    获取帖子的标题、链接信息，并从中筛选出有特定关键词的帖子
    :param html: 处理后的HTML页面
    :param m: month
    :param pn: 页码
    :return: 帖子信息
    """
    url = 'https://tieba.baidu.com'
    soup = BeautifulSoup(html, 'lxml')
    # 找到目标code标签，返回tag列表
    code = soup.find_all('code', attrs={'id': 'pagelet_html_frs-list/pagelet/thread_list'})
    # 提取code标签的内容（注释），返回列表
    comment = code[0].contents
    # print(type(comment[0]))
    # comment = code[0].string
    # print(type(comment))
    # 重新开始解析comment
    soup = BeautifulSoup(comment[0], 'lxml')
    # soup = BeautifulSoup(comment, 'lxml')

    # 找到目标li标签
    info = []

    # # 先找到置顶帖
    # litags_top = soup.find_all('li', attrs={'class': 'j_thread_list thread_top j_thread_list clearfix'})
    # for li in litags_top:
    #     info_top = dict()
    #     try:
    #         info_top['title'] = li.find('a', attrs={'class': 'j_th_tit'}).text.strip()
    #         info_top['link'] = ''.join([url, li.find('a', attrs={'class': 'j_th_tit'})['href']])
    #         info_top['time'] = li.find('span', attrs={'class': 'pull-right is_show_create_time'}).text.strip()
    #         info.append(info_top)
    #     except:
    #         print("错误：获取置顶帖标题失败！")

    # 再找到常规帖，提取标题、链接、发表日期、摘要信息
    litags = soup.find_all('li', attrs={'class': 'j_thread_list clearfix'})
    try:
        for li in litags:
            info_norm = dict()
            info_norm['title'] = li.find('a', attrs={'class': 'j_th_tit'}).text.strip()
            info_norm['link'] = ''.join([url, li.find('a', attrs={'class': 'j_th_tit'})['href']])
            info_norm['date'] = li.find('span', attrs={'class': 'pull-right is_show_create_time'}).text.strip()
            info_norm['abstract'] = li.find('div', attrs={'class': 'threadlist_abs threadlist_abs_onlyline'}). \
                text.strip()
            info.append(info_norm)
    except AttributeError as e:
        print("错误：%s" % e.args)
    except:
        print("错误：获取常规帖标题及摘要失败！")
    print('第 %s 页已经爬取成功， 开始处理...' % (pn/50+1))

    # 筛选发表日期在一个月以内，且标题和摘要里有指定关键词如['发热'，'卡'， '掉帧'， '']的帖子
    # # 获取当日日期
    # today = time.strftime('%m-%d', time.localtime(time.time()))
    # month = int(today.split('-')[0])
    # day = int(today.split('-')[1])
    #
    # if month - m >= 1:
    #     last_month = month - m
    # else:
    #     last_month = 12 + (month - m)
    # if last_month == 2 and day >= 29:
    #     one_month_before = ''.join([str(last_month), '-', '28'])
    # else:
    #     one_month_before = ''.join([str(last_month), '-', str(day)])

    # num = len(info)
    # info_new = []
    # for post in info:
    #     if ':' in post['date']:
    #         info_new.append(post)
    #     elif int(post['date'].split('-')[0]) == last_month and int(post['date'].split('-')[1]) >= day:
    #         info_new.append(post)
    #     elif int(post['date'].split('-')[0]) == month and int(post['date'].split('-')[1]) <= day:
    #         info_new.append(post)

    # 包含不同关键词的帖子分开存放
    kw_nums = len(keywords)
    try:
        info_has_kw = [[] for i in range(kw_nums)]
        for post in info:
            for i in range(kw_nums):
                if keywords[i] in post['abstract']:
                    info_has_kw[i].append(post)
                    break
        print('第 %s 页已经处理完成，开始爬取下一页...' % (pn / 50 + 1))
        # return info_has_kw
        return info_has_kw
    except IndexError as e:
        print("ERROR：%s" % e.args)


class Tool:
    # 去除img标签,7位长空格
    removeImg = re.compile('<img.*?>| {7}|')
    # 删除超链接标签
    removeAddr = re.compile('<a.*?>|</a>')
    # 把换行的标签换为\n
    replaceLine = re.compile('<tr>|<div>|</div>|</p>')
    # 将表格制表<td>替换为\t
    replaceTD= re.compile('<td>')
    # 把段落开头换为\n加空两格
    replacePara = re.compile('<p.*?>')
    # 将换行符或双换行符替换为\n
    replaceBR = re.compile('<br><br>|<br>')
    # 将其余标签剔除
    removeExtraTag = re.compile('<.*?>')

    def replace(self, x):
        x = re.sub(self.removeImg, "", x)
        x = re.sub(self.removeAddr, "", x)
        x = re.sub(self.replaceLine, "\n", x)
        x = re.sub(self.replaceTD, "\t", x)
        x = re.sub(self.replacePara, "\n    ", x)
        x = re.sub(self.replaceBR, "\n", x)
        x = re.sub(self.removeExtraTag, "", x)
        # strip()将前后多余空格删除
        return x.strip()


def get_content(info):
    """
    获取每一篇帖子的具体内容
    :param info: 帖子的基本信息{}
    :return:
    """
    try:
        url = info['link']
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) \
                                  Chrome/75.0.3770.100 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=30)
        # print(response.text)
        # html = response.text.replace('<br//>', '').replace('<br>', '').replace('<cc>', '').replace('<//cc>', '')
        # soup = BeautifulSoup(html, 'lxml')
        # print(soup)
        # # content = soup.find('div', attrs={'class': 'd_post_content j_d_post_content '}).text.strip()

        pattern = re.compile('<div id="post_content_.*?>(.*?)</div>', re.S)
        content = re.findall(pattern, response.text)
        content = Tool().replace(content[0])
        return content
    except RuntimeError as e:
        print('Runtime error: %s' % e.args)


def save2file(wb, info, keyword, savepath=os.path.dirname(os.path.realpath(__file__))+'\\post.xlsx'):
    """
    将爬取到的帖子内容写入到本地，保存到指定目录的txt文件中，保存目录默认为当前目录。
    :param info: 帖子的完整内容[]
    :param savepath: 输出文件路径，默认为当前目录
    :return:
    """
    try:
        sht = wb.create_sheet(title='%s' % keyword)
        sht['A1'] = '标题'
        sht['B1'] = '链接'
        sht['C1'] = '正文'
        num = len(info)
        for i in range(2, num+2):
            sht.cell(row=i, column=1, value=info[i-2]['title'])
            sht.cell(row=i, column=2, value=info[i-2]['link'])
            sht.cell(row=i, column=3, value=info[i-2]['content'])
        wb.save(savepath)
        # for post in info:
        # with open(savepath, 'a+', encoding='utf-8') as f:
        #     for post in info:
        #         f.write('标题：{} \n 链接：{} \n 正文：{} \n --------------------------------------------------- \n '
        #                 .format(post['title'], post['link'], post['content']))
    except IOError as e:
        print("IO error: {0}".format(e))


if __name__ == '__main__':
    post_name = '王者荣耀'
    tab = 'main'
    # keywords = ['卡顿', '卡死', '闪退', '无响应', '掉帧', '发热', '慢']
    keywords = ['交友']
    page_nums = 10
    # html = get_html(post_name=post_name, tab=tab, pn=0 * 50)
    # info = get_post_info(html, 0*50, keywords)
    # print(get_content(info[0][0]))

    # 循环控制爬取的页数
    for pn in range(page_nums):
        html = get_html(post_name=post_name, tab=tab, pn=pn*50)
        info = get_post_info(html, pn*50, keywords)
        # 循环获取每一篇帖子的正文内容
        # 获取每一篇帖子的url
        num = len(keywords)
        for i in range(num):
            res = []
            for each in info[i]:
                content = dict()
                content['content'] = get_content(each)
                content['title'] = each['title']
                content['link'] = each['link']
                res.append(content)
            save2file(res, savepath=os.path.dirname(os.path.realpath(__file__))+'\\post_%s.txt' % keywords[i])
        print("当前页面已经保存到本地！")
    print('-------所有帖子下载完成-------')
